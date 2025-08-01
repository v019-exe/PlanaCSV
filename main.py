import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import csv
from collections import defaultdict
from datetime import date, datetime
import os
import ctypes
import calendar
import requests
import threading
import tempfile
import shutil
import sys
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

RANGOS_HORARIOS = {
    '00-08': [str(h).zfill(2) for h in range(0, 8)],
    '08-09': [str(h).zfill(2) for h in range(8, 9)],
    '09-13': [str(h).zfill(2) for h in range(9, 13)],
    '13-18': [str(h).zfill(2) for h in range(13, 18)],
    '18-20': [str(h).zfill(2) for h in range(18, 20)],
    '20-22': [str(h).zfill(2) for h in range(20, 22)],
    '22-23': [str(h).zfill(2) for h in range(22, 24)]
}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_FILE_PATH = os.path.join(SCRIPT_DIR, 'data.csv')

VERSION = "3.0.0"
GITHUB_REPO = "v019-exe/PlanaCSV"
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
GITHUB_RELEASE_URL = f"https://github.com/{GITHUB_REPO}/releases/latest"

def verificar_actualizacion():
    try:
        response = requests.get(GITHUB_API_URL, timeout=5)
        if response.status_code == 200:
            data = response.json()
            ultima_version = data['tag_name'].lstrip('v')
            
            def version_to_tuple(version_str):
                try:
                    return tuple(map(int, version_str.split('.')))
                except (ValueError, AttributeError):
                    return (0, 0, 0)
            
            version_actual = version_to_tuple(VERSION)
            version_disponible = version_to_tuple(ultima_version)
            
            hay_actualizacion = version_disponible > version_actual
            return hay_actualizacion, ultima_version, data['assets'][0]['browser_download_url']
        return False, VERSION, None
    except Exception:
        return False, VERSION, None

def descargar_actualizacion(url, callback):
    try:
        temp_dir = tempfile.mkdtemp()
        temp_file = os.path.join(temp_dir, "nueva_version.exe")

        response = requests.get(url, stream=True)
        response.raise_for_status()
        with open(temp_file, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        current_exe = sys.executable
        current_exe_old = f"{current_exe}.old"
        batch_file = os.path.join(temp_dir, "actualizar.bat")

        batch_script = f"""
@echo off
echo.
echo Actualizador del Analizador Plana
echo.
echo Esperando a que la aplicacion se cierre...
timeout /t 2 /nobreak > nul

echo Reemplazando archivos...
move /Y "{current_exe}" "{current_exe_old}"
move /Y "{temp_file}" "{current_exe}"

echo Iniciando la nueva version...
start "" "{current_exe}"

echo Limpiando...
rem Espera un momento antes de borrar el archivo antiguo para evitar conflictos.
ping 127.0.0.1 -n 4 > nul
del "{current_exe_old}" > nul 2>&1

rem El siguiente comando borra este script de batch.
(goto) 2>nul & del "%~f0"
"""
        with open(batch_file, 'w', encoding='utf-8') as f:
            f.write(batch_script)

        subprocess.Popen(['cmd', '/c', batch_file], creationflags=subprocess.CREATE_NO_WINDOW)
        sys.exit(0)

    except Exception as e:
        callback(False, str(e))
        return



def obtener_fechas_semana(año, mes, semana_objetivo):
    cal = calendar.Calendar(firstweekday=0)  
    try:
        semanas_del_mes = cal.monthdayscalendar(año, mes)
    except calendar.IllegalMonthError:
        return []

    if not (1 <= semana_objetivo <= len(semanas_del_mes)):
        return []

    semana_seleccionada = semanas_del_mes[semana_objetivo - 1]
    
    fechas_formateadas = []
    for dia in semana_seleccionada:
        if dia != 0:
            fechas_formateadas.append(date(año, mes, dia).strftime('%d/%m/%Y'))
    return fechas_formateadas

def obtener_fechas_mes_completo(año, mes):
    try:
        import calendar
        num_dias = calendar.monthrange(año, mes)[1]
        
        fechas_formateadas = []
        for dia in range(1, num_dias + 1):
            fechas_formateadas.append(date(año, mes, dia).strftime('%d/%m/%Y'))
        
        return fechas_formateadas
    except (ValueError, calendar.IllegalMonthError):
        return []

def contar_llamadas_perdidas_por_rango(fechas_objetivo, archivo_csv):
    conteos = defaultdict(lambda: defaultdict(int))
    
    for fecha in fechas_objetivo:
        for rango in RANGOS_HORARIOS:
            conteos[fecha][rango] = 0

    try:
        with open(archivo_csv, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                agente = row.get('Agente', '').strip().lower()
                fecha_raw = row.get('Fecha de inicio de la llamada', '').strip()
                hora_raw = row.get('Hora de inicio de la llamada', '').strip()

                if not fecha_raw or not hora_raw:
                    continue

                try:
                    fecha_iso = fecha_raw.split(' ')[0]
                    fecha_formateada = datetime.strptime(fecha_iso, '%Y-%m-%d').strftime('%d/%m/%Y')
                    hora = hora_raw[:2]

                    
                    if agente == 'undefined' and fecha_formateada in fechas_objetivo:
                        for rango, horas_rango in RANGOS_HORARIOS.items():
                            if hora in horas_rango:
                                conteos[fecha_formateada][rango] += 1
                                break
                except (ValueError, IndexError):
                    continue
    except FileNotFoundError:
        messagebox.showerror("Error", f"No se encontró el archivo: {archivo_csv}")
        return None
    return conteos

def contar_llamadas_por_agente_y_dia(fechas_objetivo, archivo_csv):
    contadores = defaultdict(lambda: defaultdict(int))
    try:
        with open(archivo_csv, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                agente = row.get('Agente', '').strip()
                fecha_raw = row.get('Fecha de inicio de la llamada', '').strip()

                if not fecha_raw or not agente or agente.lower() == 'undefined':
                    continue
                
                try:
                    fecha_iso = fecha_raw.split(' ')[0]
                    fecha_formateada = datetime.strptime(fecha_iso, '%Y-%m-%d').strftime('%d/%m/%Y')
                    
                    if fecha_formateada in fechas_objetivo:
                        contadores[agente][fecha_formateada] += 1
                except ValueError:
                    continue
    except FileNotFoundError:
        
        return None
    return contadores

def actualizar_excel_existente(conteos_perdidas, fechas_objetivo, archivo_excel, año, mes):
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(archivo_excel)
        
        nombre_hoja = None
        año_corto = str(año)[-2:]
        
        posibles_nombres = [
            f"{año}",
            f"{año}-{mes:02d}",
            f"{año}_{mes:02d}",
            f"{año}-{mes}",
            f"{año}_{mes}",
            f"{mes:02d}-{año}",
            f"{mes:02d}_{año}",
            f"{mes}-{año}",
            f"{mes}_{año}",
            f"Mes {mes} {año}",
            f"{mes}/{año}",
            f"{mes:02d}/{año}",
            
            f"{año_corto}",
            f"{año_corto}-{mes:02d}",
            f"{año_corto}_{mes:02d}",
            f"{año_corto}-{mes}",
            f"{año_corto}_{mes}",
            f"{mes:02d}-{año_corto}",
            f"{mes:02d}_{año_corto}",
            f"{mes}-{año_corto}",
            f"{mes}_{año_corto}",
            f"Mes {mes} {año_corto}",
            f"{mes}/{año_corto}",
            f"{mes:02d}/{año_corto}",
        ]
        
        meses_nombres = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        
        mes_nombre = meses_nombres.get(mes, "")
        if mes_nombre:
            posibles_nombres.extend([
                mes_nombre,
                mes_nombre.upper(),
            ])
            
            posibles_nombres.extend([
                f"{mes_nombre} {año}",
                f"{mes_nombre}_{año}",
                f"{mes_nombre}-{año}",
                f"{año} {mes_nombre}",
                f"{año}_{mes_nombre}",
                f"{año}-{mes_nombre}",
            ])
            mes_nombre_upper = mes_nombre.upper()
            posibles_nombres.extend([
                f"{mes_nombre_upper} {año}",
                f"{mes_nombre_upper}_{año}",
                f"{mes_nombre_upper}-{año}",
                f"{año} {mes_nombre_upper}",
                f"{año}_{mes_nombre_upper}",
                f"{año}-{mes_nombre_upper}",
            ])
            
            posibles_nombres.extend([
                f"{mes_nombre} {año_corto}",
                f"{mes_nombre}_{año_corto}",
                f"{mes_nombre}-{año_corto}",
                f"{año_corto} {mes_nombre}",
                f"{año_corto}_{mes_nombre}",
                f"{año_corto}-{mes_nombre}",
            ])
            posibles_nombres.extend([
                f"{mes_nombre_upper} {año_corto}",
                f"{mes_nombre_upper}_{año_corto}",
                f"{mes_nombre_upper}-{año_corto}",
                f"{año_corto} {mes_nombre_upper}",
                f"{año_corto}_{mes_nombre_upper}",
                f"{año_corto}-{mes_nombre_upper}",
            ])
        
        posibles_nombres_upper = [nombre.upper() for nombre in posibles_nombres]
        posibles_nombres.extend(posibles_nombres_upper)
        
        palabras_extendidas = [
            'HORAS EXTENDIDAS', 'HORAS EXTEN', 'EXTENDIDAS', 'EXTEN',
            'HORAS EXT', 'H EXTENDIDAS', 'H EXTEN', 'H EXT',
            'EXTENDED', 'EXT HOURS', 'EXTENDED HOURS'
        ]
        
        def es_hoja_extendida(nombre_hoja):
            nombre_upper = nombre_hoja.upper()
            return any(palabra in nombre_upper for palabra in palabras_extendidas)
        
        hojas_candidatas = []
        for sheet_name in wb.sheetnames:
            if sheet_name in posibles_nombres and not es_hoja_extendida(sheet_name):
                hojas_candidatas.append(sheet_name)
        
        if hojas_candidatas:
            mes_nombre_upper = meses_nombres.get(mes, "").upper()
            
            for candidata in hojas_candidatas:
                if candidata.upper() == mes_nombre_upper:
                    nombre_hoja = candidata
                    break
            
            if not nombre_hoja:
                for candidata in hojas_candidatas:
                    if mes_nombre_upper in candidata.upper() and (str(año) in candidata or año_corto in candidata):
                        nombre_hoja = candidata
                        break
            
            if not nombre_hoja:
                nombre_hoja = hojas_candidatas[0]
        
        if not nombre_hoja:
            for sheet_name in wb.sheetnames:
                if es_hoja_extendida(sheet_name):
                    continue
                    
                sheet_upper = sheet_name.upper()
                if str(año) in sheet_name and (str(mes) in sheet_name or str(mes).zfill(2) in sheet_name):
                    nombre_hoja = sheet_name
                    break
                if año_corto in sheet_name and (str(mes) in sheet_name or str(mes).zfill(2) in sheet_name):
                    nombre_hoja = sheet_name
                    break
                if mes_nombre and mes_nombre.upper() in sheet_upper and str(año) in sheet_name:
                    nombre_hoja = sheet_name
                    break
                if mes_nombre and mes_nombre.upper() in sheet_upper and año_corto in sheet_name:
                    nombre_hoja = sheet_name
                    break
        
        if not nombre_hoja:
            hojas_disponibles = ", ".join(wb.sheetnames)
            raise Exception(f"No se encontró una hoja para {mes}/{año}. Hojas disponibles: {hojas_disponibles}")
        
        ws = wb[nombre_hoja]
        
        dias_semana = {
            0: 'LUNES',
            1: 'MARTES', 
            2: 'MIÉRCOLES',
            3: 'JUEVES',
            4: 'VIERNES',
            5: 'SÁBADO',
            6: 'DOMINGO'
        }
        
        rangos_orden = ['00-08', '08-09', '09-13', '13-18', '18-20', '20-22', '22-23']
        
        fechas_ordenadas = sorted(fechas_objetivo, key=lambda x: datetime.strptime(x, '%d/%m/%Y'))
        
        filas_actualizadas = 0
        for fecha in fechas_ordenadas:
            fecha_obj = datetime.strptime(fecha, '%d/%m/%Y')
            dia_semana = dias_semana[fecha_obj.weekday()]
            
            fila_encontrada = None
            for row in range(1, ws.max_row + 1):
                celda_fecha = ws.cell(row=row, column=2).value
                if celda_fecha:
                    if hasattr(celda_fecha, 'strftime'):
                        fecha_celda_str = celda_fecha.strftime('%d/%m/%Y')
                    else:
                        fecha_celda_str = str(celda_fecha).strip()
                    
                    if fecha_celda_str == fecha:
                        fila_encontrada = row
                        break
            
            if fila_encontrada:
                for col_idx, rango in enumerate(rangos_orden, 3):
                    valor = conteos_perdidas.get(fecha, {}).get(rango, 0)
                    if 3 <= col_idx <= 9:
                        ws.cell(row=fila_encontrada, column=col_idx, value=valor)
                
                filas_actualizadas += 1
        
        wb.save(archivo_excel)
        return True, f"Datos actualizados en la hoja: {nombre_hoja} ({filas_actualizadas} filas actualizadas)"
        
    except FileNotFoundError:
        raise Exception(f"No se encontró el archivo Excel: {archivo_excel}")
    except Exception as e:
        raise Exception(f"Error al actualizar el archivo Excel: {str(e)}")
        
    return False, "Error desconocido"


def actualizar_agentes_excel(conteos_agentes, fechas_objetivo, archivo_excel, año, mes):
    """
    Actualiza un archivo Excel existente con los datos de agentes
    Busca dinámicamente los nombres de agentes en las filas y actualiza sus totales diarios
    """
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(archivo_excel)
        
        nombre_hoja = None
        año_corto = str(año)[-2:]
        
        meses_nombres = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        
        mes_nombre = meses_nombres.get(mes, "")
        
        palabras_extendidas = [
            'HORAS EXTENDIDAS', 'HORAS EXTEN', 'EXTENDIDAS', 'EXTEN',
            'HORAS EXT', 'H EXTENDIDAS', 'H EXTEN', 'H EXT',
            'EXTENDED', 'EXT HOURS', 'EXTENDED HOURS'
        ]
        
        def es_hoja_extendida(nombre_hoja):
            nombre_upper = nombre_hoja.upper()
            return any(palabra in nombre_upper for palabra in palabras_extendidas)
        
        hojas_candidatas = []
        for sheet_name in wb.sheetnames:
            if not es_hoja_extendida(sheet_name):
                if (mes_nombre and mes_nombre.upper() in sheet_name.upper()) or \
                   (str(año) in sheet_name or año_corto in sheet_name):
                    hojas_candidatas.append(sheet_name)
        
        if hojas_candidatas:
            mes_nombre_upper = mes_nombre.upper() if mes_nombre else ""
            for candidata in hojas_candidatas:
                if candidata.upper() == mes_nombre_upper:
                    nombre_hoja = candidata
                    break
            
            if not nombre_hoja:
                nombre_hoja = hojas_candidatas[0]
        
        if not nombre_hoja:
            hojas_disponibles = ", ".join(wb.sheetnames)
            raise Exception(f"No se encontró una hoja para {mes}/{año}. Hojas disponibles: {hojas_disponibles}")
        
        ws = wb[nombre_hoja]
        
        def normalizar_nombre(nombre):
            import unicodedata
            import re
            
            nombre_sin_tildes = unicodedata.normalize('NFD', nombre)
            nombre_sin_tildes = ''.join(c for c in nombre_sin_tildes if unicodedata.category(c) != 'Mn')
            
            return nombre_sin_tildes.lower().strip()
        
        def extraer_nombre_del_csv(nombre_csv):
            import re
            
            match = re.search(r'\(([^)]+)\)', nombre_csv)
            if match:
                return match.group(1).strip()
            
            return nombre_csv.strip()
        
        def nombres_coinciden(nombre_csv, nombre_excel):
            nombre_limpio_csv = extraer_nombre_del_csv(nombre_csv)
            
            csv_norm = normalizar_nombre(nombre_limpio_csv)
            excel_norm = normalizar_nombre(nombre_excel)
            
            if csv_norm == excel_norm:
                return True
            
            if csv_norm in excel_norm or excel_norm in csv_norm:
                return True
            
            min_len = min(len(csv_norm), len(excel_norm))
            if min_len >= 4:
                if csv_norm[:min_len] == excel_norm[:min_len]:
                    return True
            
            palabras_excel = excel_norm.replace('.', ' ').replace('(', ' ').replace(')', ' ').split()
            for palabra in palabras_excel:
                if len(palabra) >= 3:
                    if csv_norm == palabra:
                        return True
                    if len(csv_norm) >= 4 and len(palabra) >= 4:
                        min_word_len = min(len(csv_norm), len(palabra))
                        if min_word_len >= 4 and csv_norm[:min_word_len] == palabra[:min_word_len]:
                            return True
            
            palabras_csv = csv_norm.replace('.', ' ').replace('(', ' ').replace(')', ' ').split()
            for palabra_csv in palabras_csv:
                if len(palabra_csv) >= 3:
                    for palabra_excel in palabras_excel:
                        if len(palabra_excel) >= 3:
                            if palabra_csv == palabra_excel:
                                return True
                            if len(palabra_csv) >= 4 and len(palabra_excel) >= 4:
                                min_word_len = min(len(palabra_csv), len(palabra_excel))
                                if min_word_len >= 4 and palabra_csv[:min_word_len] == palabra_excel[:min_word_len]:
                                    return True
            
            casos_especiales = {
                'consuelo': ['consu', 'ext consu', 'ext. consu'],
                'raul': ['raul', 'raúl'],
                'eliana': ['eliana'],
                'pilar': ['pilar'],
                'victoria': ['victoria'],
                'ayoub': ['ayoub']
            }
            
            for nombre_base, variaciones in casos_especiales.items():
                if csv_norm == nombre_base or any(var in csv_norm for var in variaciones):
                    if any(var in excel_norm for var in variaciones) or nombre_base in excel_norm:
                        return True
            
            return False
        
        agentes_encontrados = {}
        filas_actualizadas = 0
        
        for col in range(14, ws.max_column + 1):
            celda_agente = ws.cell(row=3, column=col).value
            if celda_agente and isinstance(celda_agente, str):
                nombre_excel = celda_agente.strip()
                
                encabezados_excluir = [
                    'LLAMADAS', 'RECIBIDAS', 'EXTENSIÓN', 'TOTAL', 'FRANJAS', 'HORARIAS', 'DIARIAS', 'ATENDIDAS',
                    'DE 00:00', 'DE 8:00', 'DE 9:00', 'DE 13:00', 'DE 18:00', 'DE 22:00', 'A PARTIR', 'SUM('
                ]
                
                if (not any(enc in nombre_excel.upper() for enc in encabezados_excluir) and
                    len(nombre_excel) > 2 and
                    not nombre_excel.startswith('=')):
                    
                    agente_encontrado = None
                    for nombre_csv in conteos_agentes.keys():
                        if nombres_coinciden(nombre_csv, nombre_excel):
                            agente_encontrado = nombre_csv
                            break
                    
                    if agente_encontrado:
                        agentes_encontrados[agente_encontrado] = col
                    else:
                        agentes_encontrados[nombre_excel] = col
        
        columnas_a_actualizar = {}
        for nombre_agente, col_agente in agentes_encontrados.items():
            columnas_a_actualizar[col_agente] = nombre_agente
        
        for col_agente, nombre_agente in columnas_a_actualizar.items():
            if nombre_agente in conteos_agentes:
                datos_agente = conteos_agentes[nombre_agente]
            else:
                datos_agente = {fecha: 0 for fecha in fechas_objetivo}
            
            for fecha in fechas_objetivo:
                total_llamadas = datos_agente.get(fecha, 0)
                
                for row in range(3, ws.max_row + 1):
                    celda_fecha = ws.cell(row=row, column=2).value
                    if celda_fecha:
                        if hasattr(celda_fecha, 'strftime'):
                            fecha_celda_str = celda_fecha.strftime('%d/%m/%Y')
                        else:
                            fecha_celda_str = str(celda_fecha).strip()
                        
                        if 'TOTAL' in fecha_celda_str.upper():
                            continue
                        
                        if fecha_celda_str == fecha:
                            ws.cell(row=row, column=col_agente, value=total_llamadas)
                            break
            
            filas_actualizadas += 1
        
        wb.save(archivo_excel)
        return True, f"Datos de agentes actualizados en la hoja: {nombre_hoja} ({filas_actualizadas} agentes actualizados)"
        
    except FileNotFoundError:
        raise Exception(f"No se encontró el archivo Excel: {archivo_excel}")
    except Exception as e:
        raise Exception(f"Error al actualizar agentes en Excel: {str(e)}")
        
    return False, "Error desconocido"




class AnalizadorCsv(tk.Tk):
    def __init__(self, default_csv_path=""):
        super().__init__()

        myappid = 'plana.analizador.llamadas.1'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

        self.title(f"Analizador de llamadas Plana v{VERSION}")
        self.geometry("500x500")
        self.iconbitmap(os.path.join(SCRIPT_DIR, 'Plana.ico'))
        
        self.csv_path_var = tk.StringVar(value=default_csv_path)
        self.excel_path_var = tk.StringVar(value="")
        self.año_var = tk.StringVar(value=date.today().year)
        self.mes_var = tk.StringVar(value=date.today().month)
        self.semana_var = tk.IntVar(value=1)
        self.mes_completo_var = tk.BooleanVar(value=False)
        
        self.style = ttk.Style()
        self.style.theme_use('clam')  
        
        self.configure(bg="#f0f0f0")
        self.style.configure('TFrame', background="#f0f0f0")
        self.style.configure('TLabel', background="#f0f0f0", font=('Segoe UI', 10))
        self.style.configure('TButton', font=('Segoe UI', 10))
        self.style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'))
        
        self._crear_widgets()
        
        self.after(1000, self.verificar_actualizaciones_silencioso)

    def _toggle_semana_selector(self):
        if self.mes_completo_var.get():
            self.semana_combo.config(state="disabled")
            self.semana_label.config(foreground="gray")
        else:
            self.semana_combo.config(state="readonly")
            self.semana_label.config(foreground="black")

    def _crear_widgets(self):
        
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        main_frame = ttk.Frame(self.notebook, padding=15)
        self.notebook.add(main_frame, text="Análisis")
        
        main_frame.grid_columnconfigure(1, weight=1)
        
        titulo_label = ttk.Label(main_frame, text="Analizador de llamadas Plana", 
                               font=('Segoe UI', 14, 'bold'))
        titulo_label.grid(row=0, column=0, columnspan=3, pady=(0, 15), sticky="w")
        
        param_frame = ttk.LabelFrame(main_frame, text="Parámetros", padding=10)
        param_frame.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(0, 15))
        param_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Label(param_frame, text="Año:").grid(row=0, column=0, sticky="w", pady=5, padx=5)
        ttk.Entry(param_frame, textvariable=self.año_var, width=10).grid(row=0, column=1, sticky="w", pady=5)
        
        ttk.Label(param_frame, text="Mes:").grid(row=1, column=0, sticky="w", pady=5, padx=5)
        ttk.Entry(param_frame, textvariable=self.mes_var, width=10).grid(row=1, column=1, sticky="w", pady=5)
        
        mes_completo_check = ttk.Checkbutton(param_frame, text="Rellenar mes completo", 
                                           variable=self.mes_completo_var, 
                                           command=self._toggle_semana_selector)
        mes_completo_check.grid(row=2, column=0, columnspan=2, sticky="w", pady=5, padx=5)
        
        self.semana_label = ttk.Label(param_frame, text="Semana:")
        self.semana_label.grid(row=3, column=0, sticky="w", pady=5, padx=5)
        semanas_opciones = [1, 2, 3, 4, 5, 6]
        self.semana_combo = ttk.Combobox(param_frame, textvariable=self.semana_var, values=semanas_opciones, width=8, state="readonly")
        self.semana_combo.grid(row=3, column=1, sticky="w", pady=5)
        self.semana_combo.current(0)
        
        file_frame = ttk.LabelFrame(main_frame, text="Archivo CSV de datos", padding=10)
        file_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 15))
        file_frame.grid_columnconfigure(0, weight=1)
        
        entry_path = ttk.Entry(file_frame, textvariable=self.csv_path_var)
        entry_path.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ttk.Button(file_frame, text="Buscar...", command=self._seleccionar_archivo).grid(row=0, column=1)
        
        excel_frame = ttk.LabelFrame(main_frame, text="Archivo Excel de destino", padding=10)
        excel_frame.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(0, 15))
        excel_frame.grid_columnconfigure(0, weight=1)
        
        entry_excel = ttk.Entry(excel_frame, textvariable=self.excel_path_var)
        entry_excel.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ttk.Button(excel_frame, text="Buscar...", command=self._seleccionar_excel).grid(row=0, column=1)
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=4, column=0, columnspan=3, sticky="ew")
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Button(btn_frame, text="Analizar", command=self.realizar_analisis, 
                  style='Accent.TButton', width=20).grid(row=0, column=0, pady=10, padx=(0, 5))
        
        ttk.Button(btn_frame, text="Actualizar Excel", command=self.exportar_excel, 
                  width=20).grid(row=0, column=1, pady=10, padx=(5, 0))
        
        update_frame = ttk.Frame(self.notebook, padding=15)
        self.notebook.add(update_frame, text="Actualizaciones")
        
        update_frame.grid_columnconfigure(0, weight=1)
        
        ttk.Label(update_frame, text=f"Versión actual: {VERSION}", 
                font=('Segoe UI', 11)).grid(row=0, column=0, sticky="w", pady=(0, 10))
        
        ttk.Button(update_frame, text="Verificar actualizaciones", 
                  command=self.verificar_actualizaciones_manual).grid(row=1, column=0, sticky="w", pady=5)
        
        self.update_status = ttk.Label(update_frame, text="")
        self.update_status.grid(row=2, column=0, sticky="w", pady=5)
        
        status_frame = ttk.Frame(self)
        status_frame.pack(fill="x", side="bottom", padx=10, pady=5)
        
        self.status_label = ttk.Label(status_frame, text=f"Analizador de llamadas Plana v{VERSION}", anchor="w")
        self.status_label.pack(side="left")
        
        self.update_button = ttk.Button(status_frame, text="¡Nueva versión disponible!", 
                                      command=self.mostrar_dialogo_actualizacion)
        

    def _seleccionar_archivo(self):
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo CSV",
            filetypes=(("Archivos CSV", "*.csv"), ("Todos los archivos", "*.*"))
        )
        if filepath:  
            self.csv_path_var.set(filepath)

    def _seleccionar_excel(self):
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=(("Archivos Excel", "*.xlsx"), ("Archivos Excel", "*.xls"), ("Todos los archivos", "*.*"))
        )
        if filepath:  
            self.excel_path_var.set(filepath)

    def _mostrar_resultados(self, texto_resultados):
        ventana_resultados = tk.Toplevel(self)
        ventana_resultados.title("Resultados del Análisis")
        ventana_resultados.geometry("750x550")
        
        ventana_resultados.transient(self)
        ventana_resultados.grab_set()
        
        try:
            ventana_resultados.iconbitmap(os.path.join(SCRIPT_DIR, 'Plana.ico'))
        except tk.TclError:
            pass
        
        main_frame = ttk.Frame(ventana_resultados, padding=15)
        main_frame.pack(fill="both", expand=True)
        
        ttk.Label(main_frame, text="Resultados del Análisis", 
                font=('Segoe UI', 14, 'bold')).pack(anchor="w", pady=(0, 15))
        
        text_frame = ttk.LabelFrame(main_frame, text="Informe detallado")
        text_frame.pack(fill="both", expand=True)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 10), 
                            bg="#f9f9f9", padx=10, pady=10)
        scrollbar = ttk.Scrollbar(text_frame, command=text_widget.yview)
        text_widget.config(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        text_widget.insert(tk.END, texto_resultados)
        
        texto = texto_resultados.split('\n')
        pos = "1.0"
        for i, linea in enumerate(texto):
            if "===" in linea or linea.startswith("---"):
                text_widget.tag_add(f"titulo{i}", pos, f"{pos.split('.')[0]}.end")
                text_widget.tag_config(f"titulo{i}", font=("Consolas", 10, "bold"), foreground="#0066cc")
            elif linea.startswith("Resultados para"):
                text_widget.tag_add(f"titulo{i}", pos, f"{pos.split('.')[0]}.end")
                text_widget.tag_config(f"titulo{i}", font=("Consolas", 12, "bold"))
            elif linea.startswith("Agente:") or linea.startswith("Día:"):
                text_widget.tag_add(f"subtitulo{i}", pos, f"{pos.split('.')[0]}.end")
                text_widget.tag_config(f"subtitulo{i}", font=("Consolas", 10, "bold"))
            elif "TOTAL" in linea:
                text_widget.tag_add(f"total{i}", pos, f"{pos.split('.')[0]}.end")
                text_widget.tag_config(f"total{i}", font=("Consolas", 10, "bold"), foreground="#006600")
            
            pos = f"{int(pos.split('.')[0]) + 1}.0"
        
        text_widget.config(state='disabled')
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=(15, 0))
        
        ttk.Button(btn_frame, text="Cerrar", 
                 command=ventana_resultados.destroy).pack(side="right")

    def verificar_actualizaciones_silencioso(self):
        def check_bg():
            hay_actualizacion, nueva_version, download_url = verificar_actualizacion()
            if hay_actualizacion:
                self.after(0, lambda: self.mostrar_notificacion_actualizacion(nueva_version, download_url))
        
        threading.Thread(target=check_bg, daemon=True).start()
    
    def verificar_actualizaciones_manual(self):
        self.update_status.config(text="Verificando actualizaciones...")
        
        def check_bg():
            hay_actualizacion, nueva_version, download_url = verificar_actualizacion()
            if hay_actualizacion:
                self.after(0, lambda: self.mostrar_dialogo_actualizacion(nueva_version, download_url))
            else:
                self.after(0, lambda: self.update_status.config(
                    text=f"Estás utilizando la última versión ({VERSION})"))
        
        threading.Thread(target=check_bg, daemon=True).start()
    
    def mostrar_notificacion_actualizacion(self, nueva_version, download_url):
        self.nueva_version = nueva_version
        self.download_url = download_url
        self.update_button.pack(side="right", padx=5)
    
    def mostrar_dialogo_actualizacion(self, nueva_version=None, download_url=None):
        if nueva_version is None:
            nueva_version = self.nueva_version
            download_url = self.download_url
            
        respuesta = messagebox.askyesno(
            "Nueva versión disponible",
            f"Hay una nueva versión disponible: v{nueva_version}\n\n"
            f"Tu versión actual es: v{VERSION}\n\n"
            "¿Deseas descargar e instalar la actualización ahora?\n"
            "(La aplicación se reiniciará automáticamente)"
        )
        
        if respuesta:
            self.update_status.config(text="Descargando actualización...")
            
            def on_complete(exito, mensaje):
                if not exito:
                    self.after(0, lambda: messagebox.showerror(
                        "Error de actualización", 
                        f"No se pudo completar la actualización: {mensaje}"))
                    self.after(0, lambda: self.update_status.config(
                        text="Error al actualizar. Intente más tarde."))
            
            threading.Thread(
                target=descargar_actualizacion,
                args=(download_url, on_complete),
                daemon=True
            ).start()
    
    def realizar_analisis(self):
        csv_path = self.csv_path_var.get()
        if not csv_path or not os.path.exists(csv_path):
            messagebox.showerror("Error de Archivo", "La ruta del archivo CSV no es válida o está vacía. Por favor, selecciona un archivo.")
            return

        try:
            año = int(self.año_var.get())
            mes = int(self.mes_var.get())
            semana = self.semana_var.get()
        except ValueError:
            messagebox.showerror("Error de Entrada", "El año y el mes deben ser números.")
            return

        if self.mes_completo_var.get():
            dias_objetivo = obtener_fechas_mes_completo(año, mes)
            periodo_texto = f"Mes Completo {mes}, Año {año}"
        else:
            dias_objetivo = obtener_fechas_semana(año, mes, semana)
            periodo_texto = f"Semana {semana} del Mes {mes}, Año {año}"
        
        if not dias_objetivo:
            if self.mes_completo_var.get():
                messagebox.showwarning("Aviso", f"No se encontraron días para el mes {mes}/{año}.")
            else:
                messagebox.showwarning("Aviso", f"No se encontraron días para la semana {semana} del mes {mes}/{año}.")
            return

        conteos_perdidas = contar_llamadas_perdidas_por_rango(dias_objetivo, csv_path)
        conteos_agentes = contar_llamadas_por_agente_y_dia(dias_objetivo, csv_path)

        if conteos_perdidas is None or conteos_agentes is None:
            return 

        texto_final = f"Resultados para {periodo_texto}\n"
        texto_final += "=" * 50 + "\n\n"

        texto_final += "--- LLAMADAS PERDIDAS POR FRANJA HORARIA ---\n"
        for dia in dias_objetivo:
            texto_final += f'\nDía: {dia}\n'
            total_dia = sum(conteos_perdidas[dia].values())
            for rango in RANGOS_HORARIOS:
                cantidad = conteos_perdidas[dia][rango]
                texto_final += f'  - Rango {rango}: {cantidad} llamadas\n'
            texto_final += f'  TOTAL DÍA: {total_dia} llamadas perdidas\n'

        texto_final += "\n" + "=" * 50 + "\n\n"
        texto_final += "--- LLAMADAS ATENDIDAS POR AGENTE ---\n"
        for agente, fechas_dict in sorted(conteos_agentes.items()):
            texto_final += f'\nAgente: {agente}\n'
            for fecha in dias_objetivo:
                llamadas_dia = fechas_dict.get(fecha, 0)
                if llamadas_dia > 0:
                    texto_final += f'  - {fecha}: {llamadas_dia} llamadas\n'
            total_agente = sum(fechas_dict.values())
            texto_final += f'  TOTAL AGENTE: {total_agente} llamadas\n'

        self._mostrar_resultados(texto_final)
    
    def exportar_excel(self):
        csv_path = self.csv_path_var.get()
        if not csv_path or not os.path.exists(csv_path):
            messagebox.showerror("Error de Archivo", "La ruta del archivo CSV no es válida o está vacía. Por favor, selecciona un archivo.")
            return

        excel_path = self.excel_path_var.get()
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showerror("Error de Archivo", "La ruta del archivo Excel no es válida o está vacía. Por favor, selecciona un archivo Excel.")
            return

        try:
            año = int(self.año_var.get())
            mes = int(self.mes_var.get())
            semana = self.semana_var.get()
        except ValueError:
            messagebox.showerror("Error de Entrada", "El año y el mes deben ser números.")
            return

        if self.mes_completo_var.get():
            dias_objetivo = obtener_fechas_mes_completo(año, mes)
            if not dias_objetivo:
                messagebox.showwarning("Aviso", f"No se encontraron días para el mes {mes}/{año}.")
                return
        else:
            dias_objetivo = obtener_fechas_semana(año, mes, semana)
            if not dias_objetivo:
                messagebox.showwarning("Aviso", f"No se encontraron días para la semana {semana} del mes {mes}/{año}.")
                return

        conteos_perdidas = contar_llamadas_perdidas_por_rango(dias_objetivo, csv_path)
        if conteos_perdidas is None:
            return
        
        conteos_agentes = contar_llamadas_por_agente_y_dia(dias_objetivo, csv_path)
        if conteos_agentes is None:
            return

        try:
            resultado1, mensaje1 = actualizar_excel_existente(conteos_perdidas, dias_objetivo, excel_path, año, mes)
            
            resultado2, mensaje2 = actualizar_agentes_excel(conteos_agentes, dias_objetivo, excel_path, año, mes)
            
            if resultado1 and resultado2:
                messagebox.showinfo("Éxito", f"Archivo Excel actualizado correctamente:\n{excel_path}\n\n{mensaje1}\n{mensaje2}")
            elif resultado1:
                messagebox.showwarning("Parcial", f"Solo se actualizaron las llamadas perdidas:\n{excel_path}\n\n{mensaje1}\n\nError en agentes: {mensaje2}")
            elif resultado2:
                messagebox.showwarning("Parcial", f"Solo se actualizaron los agentes:\n{excel_path}\n\n{mensaje2}\n\nError en llamadas perdidas: {mensaje1}")
            else:
                messagebox.showerror("Error", f"No se pudo actualizar el archivo Excel.\n{mensaje1}\n{mensaje2}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al actualizar Excel: {str(e)}")


if __name__ == "__main__":
    app = AnalizadorCsv(default_csv_path=CSV_FILE_PATH)
    app.mainloop()
